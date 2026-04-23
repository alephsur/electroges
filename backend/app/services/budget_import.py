"""CSV/XLSX budget line import.

Expected columns (header row, case-insensitive, accepts Spanish aliases):
  section      → section / capitulo / capítulo / chapter       (optional)
  type         → type / tipo  (labor|mano de obra|material|other|partida)  (required)
  description  → description / descripcion / descripción       (required)
  quantity     → quantity / cantidad / cant                    (required)
  unit         → unit / ud / unidad                            (optional)
  unit_price   → unit_price / precio / p_venta / p.unitario    (required)
  unit_cost    → unit_cost / coste                             (optional, default 0)
  discount     → discount / dto / descuento                    (optional, default 0)
"""
from __future__ import annotations

import csv
import io
import logging
import uuid
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.budget import Budget, BudgetLine, BudgetLineType, BudgetSection, BudgetStatus
from app.repositories.budget import (
    BudgetLineRepository,
    BudgetRepository,
    BudgetSectionRepository,
)

logger = logging.getLogger(__name__)


class ImportLineRow(BaseModel):
    section: str | None = None
    line_type: str  # "labor" | "material" | "other"
    description: str
    quantity: float
    unit: str | None = None
    unit_price: float
    unit_cost: float = 0.0
    line_discount_pct: float = 0.0


class ImportError(BaseModel):
    row_number: int
    field: str | None = None
    message: str


class ImportPreview(BaseModel):
    """Parsed preview (never writes to DB). Frontend calls /confirm to commit."""

    valid_rows: list[ImportLineRow]
    errors: list[ImportError]
    sections_detected: list[str]
    total_rows: int


# ── Header aliases ────────────────────────────────────────────────────────────

_SECTION_ALIASES = {"section", "capitulo", "capítulo", "chapter", "seccion", "sección"}
_TYPE_ALIASES = {"type", "tipo"}
_DESC_ALIASES = {"description", "descripcion", "descripción", "desc"}
_QTY_ALIASES = {"quantity", "cantidad", "cant", "qty"}
_UNIT_ALIASES = {"unit", "ud", "unidad", "uds"}
_PRICE_ALIASES = {
    "unit_price",
    "precio",
    "p_venta",
    "p.unitario",
    "precio_unitario",
    "precio unitario",
    "p.venta",
}
_COST_ALIASES = {"unit_cost", "coste", "costo", "cost"}
_DISCOUNT_ALIASES = {
    "discount",
    "dto",
    "descuento",
    "line_discount_pct",
    "dto %",
    "dto%",
}

_TYPE_MAP = {
    "labor": "labor",
    "mano de obra": "labor",
    "mano_de_obra": "labor",
    "trabajo": "labor",
    "material": "material",
    "materiales": "material",
    "other": "other",
    "otro": "other",
    "partida": "other",
}


def _normalize(s: str) -> str:
    return s.strip().lower()


def _match_column(header: str, aliases: set[str]) -> bool:
    return _normalize(header) in aliases


def _map_columns(headers: list[str]) -> dict[str, int]:
    """Returns {logical_field_name: column_index}."""
    mapping: dict[str, int] = {}
    for idx, h in enumerate(headers):
        h_norm = _normalize(h)
        if h_norm in _SECTION_ALIASES:
            mapping["section"] = idx
        elif h_norm in _TYPE_ALIASES:
            mapping["type"] = idx
        elif h_norm in _DESC_ALIASES:
            mapping["description"] = idx
        elif h_norm in _QTY_ALIASES:
            mapping["quantity"] = idx
        elif h_norm in _UNIT_ALIASES:
            mapping["unit"] = idx
        elif h_norm in _PRICE_ALIASES:
            mapping["unit_price"] = idx
        elif h_norm in _COST_ALIASES:
            mapping["unit_cost"] = idx
        elif h_norm in _DISCOUNT_ALIASES:
            mapping["discount"] = idx
    return mapping


def _parse_decimal(value: str | None) -> Decimal:
    if value is None:
        return Decimal("0")
    s = str(value).strip().replace(",", ".")
    if not s:
        return Decimal("0")
    return Decimal(s)


def _parse_rows(
    raw_rows: list[list[str]],
) -> tuple[list[ImportLineRow], list[ImportError], list[str]]:
    if not raw_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El fichero está vacío",
        )

    headers = raw_rows[0]
    mapping = _map_columns(headers)

    missing: list[str] = []
    for required in ("type", "description", "quantity", "unit_price"):
        if required not in mapping:
            missing.append(required)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Faltan columnas obligatorias: "
                + ", ".join(missing)
                + ". Columnas detectadas: "
                + ", ".join(headers)
            ),
        )

    valid: list[ImportLineRow] = []
    errors: list[ImportError] = []
    sections_seen: list[str] = []

    for row_idx, row in enumerate(raw_rows[1:], start=2):
        if not any((c or "").strip() for c in row):
            continue  # skip empty lines

        def get(field: str) -> str | None:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None:
                return None
            s = str(val).strip()
            return s or None

        section_name = get("section")
        if section_name and section_name not in sections_seen:
            sections_seen.append(section_name)

        raw_type = get("type") or ""
        line_type = _TYPE_MAP.get(_normalize(raw_type))
        if line_type is None:
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="type",
                    message=f"Tipo de línea inválido: '{raw_type}'. Valores válidos: labor, material, other.",
                )
            )
            continue

        description = get("description")
        if not description:
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="description",
                    message="La descripción es obligatoria",
                )
            )
            continue

        try:
            quantity = _parse_decimal(get("quantity"))
            if quantity <= 0:
                raise InvalidOperation("quantity must be > 0")
        except (InvalidOperation, ValueError):
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="quantity",
                    message=f"Cantidad inválida: '{get('quantity')}'",
                )
            )
            continue

        try:
            unit_price = _parse_decimal(get("unit_price"))
        except (InvalidOperation, ValueError):
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="unit_price",
                    message=f"Precio inválido: '{get('unit_price')}'",
                )
            )
            continue

        try:
            unit_cost = _parse_decimal(get("unit_cost"))
        except (InvalidOperation, ValueError):
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="unit_cost",
                    message=f"Coste inválido: '{get('unit_cost')}'",
                )
            )
            continue

        try:
            discount = _parse_decimal(get("discount"))
            if discount < 0 or discount > 100:
                raise InvalidOperation("discount out of range")
        except (InvalidOperation, ValueError):
            errors.append(
                ImportError(
                    row_number=row_idx,
                    field="discount",
                    message=f"Descuento inválido: '{get('discount')}'",
                )
            )
            continue

        valid.append(
            ImportLineRow(
                section=section_name,
                line_type=line_type,
                description=description,
                quantity=float(quantity),
                unit=get("unit"),
                unit_price=float(unit_price),
                unit_cost=float(unit_cost),
                line_discount_pct=float(discount),
            )
        )

    return valid, errors, sections_seen


# ── Public API ────────────────────────────────────────────────────────────────

async def parse_upload(file: UploadFile) -> list[list[str]]:
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv") or (file.content_type or "").startswith("text/"):
        # CSV: detect delimiter
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        return [list(r) for r in reader]

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="openpyxl no está instalado en el servidor",
            ) from exc
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        return rows

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Formato no soportado. Usa CSV o XLSX.",
    )


class BudgetImportService:
    def __init__(self, session: AsyncSession, tenant_id: uuid.UUID):
        self._session = session
        self._tenant_id = tenant_id
        self._budget_repo = BudgetRepository(session, tenant_id)
        self._line_repo = BudgetLineRepository(session, tenant_id)
        self._section_repo = BudgetSectionRepository(session, tenant_id)

    async def _get_draft_budget(self, budget_id: uuid.UUID) -> Budget:
        budget = await self._budget_repo.get_with_full_detail(budget_id)
        if not budget:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Presupuesto no encontrado",
            )
        if budget.status != BudgetStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Solo se pueden importar líneas en presupuestos en borrador",
            )
        return budget

    async def preview_import(
        self, budget_id: uuid.UUID, file: UploadFile
    ) -> ImportPreview:
        await self._get_draft_budget(budget_id)
        raw_rows = await parse_upload(file)
        valid, errors, sections = _parse_rows(raw_rows)
        return ImportPreview(
            valid_rows=valid,
            errors=errors,
            sections_detected=sections,
            total_rows=max(0, len(raw_rows) - 1),
        )

    async def confirm_import(
        self,
        budget_id: uuid.UUID,
        rows: list[ImportLineRow],
    ) -> dict:
        """Create budget lines (and any new sections) from the preview payload."""
        budget = await self._get_draft_budget(budget_id)

        # Existing section name → id map
        existing_sections = {s.name.strip().lower(): s for s in budget.sections}

        section_offset = max(
            (s.sort_order for s in budget.sections), default=-1
        ) + 1
        line_offset = max((l.sort_order for l in budget.lines), default=-1) + 1

        created_sections = 0
        created_section_ids: dict[str, uuid.UUID] = {
            k: v.id for k, v in existing_sections.items()
        }

        # Pre-create new sections in order of first appearance
        seen: list[str] = []
        for row in rows:
            if row.section:
                key = row.section.strip().lower()
                if key and key not in created_section_ids and key not in seen:
                    seen.append(key)

        for idx, key in enumerate(seen):
            original_name = next(
                (r.section for r in rows if r.section and r.section.strip().lower() == key),
                key,
            )
            new_section = BudgetSection(
                budget_id=budget.id,
                name=original_name,
                sort_order=section_offset + idx,
            )
            new_section = await self._section_repo.create(new_section)
            created_section_ids[key] = new_section.id
            created_sections += 1

        # Create lines
        for idx, row in enumerate(rows):
            target_section_id: uuid.UUID | None = None
            if row.section:
                target_section_id = created_section_ids.get(
                    row.section.strip().lower()
                )
            line = BudgetLine(
                budget_id=budget.id,
                section_id=target_section_id,
                line_type=BudgetLineType(row.line_type),
                sort_order=line_offset + idx,
                description=row.description,
                quantity=Decimal(str(row.quantity)),
                unit=row.unit,
                unit_price=Decimal(str(row.unit_price)),
                unit_cost=Decimal(str(row.unit_cost)),
                line_discount_pct=Decimal(str(row.line_discount_pct)),
            )
            await self._line_repo.create(line)

        await self._session.commit()
        logger.info(
            "Imported %d lines and %d new sections into budget %s",
            len(rows),
            created_sections,
            budget_id,
        )
        return {
            "imported_lines": len(rows),
            "created_sections": created_sections,
        }
